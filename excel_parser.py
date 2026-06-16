"""
Excel parser: lecture du classeur, détection des zones de données,
plages nommées, ListObjects et construction du modèle de données interne.

Principe de formatage :
  - Tous les nombres sont pris tels quels, sans détection automatique de pourcentage.
  - Les entiers non-années sont formatés avec séparateur de milliers.
  - Les flottants sont arrondis à 3 décimales max (zéros de queue supprimés),
    virgule comme séparateur décimal, espace fine comme séparateur des milliers.
  - La valeur 0 entier → chaîne vide (meilleur affichage).
  - Les années (1900-2100) sont renvoyées telles quelles (int), sans séparateur.
  - Le formatage pourcentage est laissé à l'utilisateur via les RowStyle (tab_line).

Nommage des feuilles / pages :
  - SheetModel.name est le nom original de la feuille Excel. C'est la clé technique
    utilisée pour référencer page_params, table_styles, charts, sheet_settings, etc.
    Elle ne doit jamais changer, même si l'ordre des pages est modifié.
  - SheetModel.display_name est le nom affiché dans le PDF (titre de page + sommaire).
    Si None, on retombe sur SheetModel.name. Modifiable par l'utilisateur dans
    l'onglet Document, et persisté via GlobalParams.sheet_settings (cf pdf_generator).
"""
import json
import re
from dataclasses import dataclass, field, asdict
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from util import is_year_like


# ── Dataclasses modèle ─────────────────────────────────────────────────────

@dataclass
class CellRange:
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    def to_a1(self) -> str:
        return (
            f"{get_column_letter(self.min_col)}{self.min_row}:"
            f"{get_column_letter(self.max_col)}{self.max_row}"
        )


@dataclass
class TableZone:
    name: str
    range: CellRange
    source: str   # "list_object" | "named_range" | "heuristic" | "manual"
    headers: list[str] = field(default_factory=list)
    data: list[list[Any]] = field(default_factory=list)


@dataclass
class SheetModel:
    name: str                       # nom ORIGINAL Excel — clé technique stable, ne pas modifier
    index: int
    tables: list[TableZone] = field(default_factory=list)
    named_ranges: dict[str, str] = field(default_factory=dict)
    raw_cells: dict[str, Any] = field(default_factory=dict)
    dimensions: tuple[int, int] = (0, 0)
    include: bool = True
    page_order: int = 0
    display_name: str | None = None  # nom affiché (titre de page + sommaire) ; None → name

    def get_display_name(self) -> str:
        return self.display_name if self.display_name else self.name


@dataclass
class WorkbookModel:
    file_path: str
    sheets: list[SheetModel] = field(default_factory=list)
    mapping_path: str | None = None

    def to_dict(self) -> dict:
        return asdict(self)


# ── Helpers généraux ───────────────────────────────────────────────────────

def _parse_ref(ref: str) -> CellRange | None:
    """Parse 'A1:C10' ou 'Sheet!A1:C10' → CellRange."""
    try:
        if "!" in ref:
            ref = ref.split("!", 1)[1]
        if ":" in ref:
            start, end = ref.split(":")
        else:
            start = end = ref
        start = start.replace("$", "")
        end   = end.replace("$", "")
        c1 = "".join(c for c in start if c.isalpha())
        r1 = int("".join(c for c in start if c.isdigit()))
        c2 = "".join(c for c in end if c.isalpha())
        r2 = int("".join(c for c in end if c.isdigit()))
        return CellRange(r1, column_index_from_string(c1), r2, column_index_from_string(c2))
    except Exception:
        return None


def _is_empty_value(value: Any) -> bool:
    return value is None or value == "" or _is_excel_error(value)

_EXCEL_ERRORS = frozenset({
    "#REF!", "#DIV/0!", "#N/A", "#N/A!", "#VALUE!",
    "#NAME?", "#NULL!", "#NUM!", "#ERROR!", "#GETTING_DATA",
})

def _is_excel_error(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    s = value.strip()
    return s in _EXCEL_ERRORS or (s.startswith("#") and len(s) <= 20)

def _is_numeric_non_year(value: Any) -> bool:
    if isinstance(value, bool) or _is_empty_value(value) or is_year_like(value):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        s = value.strip().replace(" ", "")
        if "," in s and "." not in s:
            s = s.replace(",", "")
        try:
            float(s)
            return not is_year_like(s)
        except ValueError:
            return False
    return False


# ── Formatage numérique ─────────────────────────────────────────────────────

def _fmt_number(num: float) -> str:
    """
    Formate un float en chaîne française avec séparateur des milliers.
    Règles :
      • Arrondi à 3 décimales max, zéros de queue supprimés.
      • Séparateur des milliers : espace sur la partie entière.
      • Séparateur décimal : virgule.
      • 0.0 → chaîne vide.
    """
    if num == 0.0:
        return ""

    sign = "-" if num < 0 else ""
    abs_num = abs(num)
    rounded = round(abs_num, 3)
    int_part = int(rounded)
    dec_val  = rounded - int_part

    int_str = f"{int_part:,}".replace(",", " ")

    if abs(dec_val) < 1e-9:
        return f"{sign}{int_str}"

    dec_str = f"{rounded:.3f}".split(".")[1].rstrip("0")
    return f"{sign}{int_str},{dec_str}"


def _format_numeric_cell(v: Any, is_mapping: bool = False) -> Any:
    """
    Formate une valeur de cellule Excel au format français.
    - Années → int (inchangées, sans séparateur).
    - Entiers non-années → chaîne avec séparateur des milliers (0 → "").
    - Flottants → chaîne avec virgule et séparateur des milliers.
    - Chaînes numériques → parsées et reformatées.
    - Autres → inchangées.
    """
    if v is None and not is_mapping:
        return None
    if isinstance(v, bool):
        return v
    if _is_excel_error(v):
        return None
    if isinstance(v, int):
        if is_year_like(v):
            return v
        return _fmt_number(float(v))

    if isinstance(v, float):
        if is_year_like(v):
            return int(v)
        return _fmt_number(v)

    if isinstance(v, str):
        s = v.strip()
        s_parse = s.replace(" ", "").replace("\xa0", "").replace(" ", "")
        if "," in s_parse and "." not in s_parse:
            s_parse = s_parse.replace(",", ".")
        try:
            f = float(s_parse)
            if is_year_like(f):
                return int(f)
            return _fmt_number(f)
        except ValueError:
            return v

    return v


# ── Extraction des données d'un tableau ────────────────────────────────────

def _extract_table_data(
    ws, cr: CellRange, is_mapping: bool = False
) -> tuple[list[str], list[list[Any]]]:
    """
    Extrait les en-têtes et les lignes de données depuis une plage de cellules Excel.

    - Supprime les colonnes d'années antérieures à start_yr.
    - Supprime les colonnes "Cumul XX" dont l'année est antérieure à start_yr.
    - Supprime les colonnes entièrement vides (sauf si en-tête numérique ou année).
    - Formate tous les nombres en français (pas de détection de pourcentage —
      c'est le rôle de RowStyle dans tab_line).
    """
    rows = list(ws.iter_rows(
        min_row=cr.min_row, max_row=cr.max_row,
        min_col=cr.min_col, max_col=cr.max_col,
        values_only=True,
    ))
    if not rows:
        return [], []

    headers = [str(h) if not _is_empty_value(h) else "" for h in rows[0]]
    data    = [list(r) for r in rows[1:]]

    try:
        from state_manager import state_manager
        start_yr = state_manager.get_start_year()
    except Exception:
        start_yr = 2025

    def _representative(idx: int) -> list:
        if idx >= len(rows):
            return []
        row = rows[idx]
        return list(row) if not all(_is_empty_value(v) for v in row) else []

    second_row = _representative(1) or _representative(2)
    third_row  = (
        _representative(2)
        if second_row != (list(rows[2]) if len(rows) > 2 else [])
        else (_representative(3) if len(rows) > 3 else [])
    )
    candidate_rows = [r for r in [second_row, third_row] if r]

    def _drop_by_year(idx: int) -> bool:
        if start_yr is None:
            return False
        candidates = [rows[0][idx]] + [r[idx] for r in candidate_rows if idx < len(r)]
        for val in candidates:
            if is_year_like(val):
                try:
                    if int(float(str(val).strip())) < start_yr:
                        return True
                except Exception:
                    pass
        return False

    def _drop_by_cumul(idx: int) -> bool:
        if start_yr is None:
            return False
        candidates = [rows[0][idx]] + [r[idx] for r in candidate_rows if idx < len(r)]
        for val in candidates:
            if not isinstance(val, str):
                continue
            vs = val.strip()
            vl = vs.lower()
            if vl.startswith("cumul"):
                if len(vs) >= 2 and vs[-2:].isdigit():
                    try:
                        if int(vs[-2:]) + 2000 < start_yr:
                            return True
                    except Exception:
                        pass
                m = re.search(r'(\d{2})/(\d{2})$', vs)
                if m:
                    try:
                        if int(m.group(2)) + 2000 < start_yr:
                            return True
                    except Exception:
                        pass
        return False

    keep: list[int] = []
    for idx, header in enumerate(rows[0]):
        col_vals = [r[idx] for r in data if idx < len(r)]
        has_data = any(not _is_empty_value(v) for v in col_vals)
        if _drop_by_year(idx) or _drop_by_cumul(idx):
            continue
        if has_data or _is_numeric_non_year(header) or is_year_like(header):
            keep.append(idx)

    if not keep:
        return [], []

    filtered_headers = [headers[i] for i in keep]

    filtered_data = []
    for row in data:
        formatted = [
            _format_numeric_cell(row[i] if i < len(row) else None, is_mapping)
            for i in keep
        ]
        filtered_data.append(formatted)

    filtered_data = [
        row for row in filtered_data
        if any(not _is_empty_value(v) for v in row)
    ]
    return filtered_headers, filtered_data


# ── Détection heuristique ──────────────────────────────────────────────────

def _detect_heuristic_tables(ws) -> list[CellRange]:
    if ws.max_row is None or ws.max_column is None:
        return []

    visited: set[tuple[int, int]] = set()
    blocks:  list[CellRange]      = []

    for row in ws.iter_rows():
        for cell in row:
            if _is_empty_value(cell.value) or (cell.row, cell.column) in visited:
                continue
            min_r = max_r = cell.row
            min_c = max_c = cell.column
            queue = [(cell.row, cell.column)]
            while queue:
                r, c = queue.pop()
                if (r, c) in visited:
                    continue
                v = ws.cell(row=r, column=c).value
                if _is_empty_value(v):
                    continue
                visited.add((r, c))
                min_r, max_r = min(min_r, r), max(max_r, r)
                min_c, max_c = min(min_c, c), max(max_c, c)
                for dr, dc in [(a, b) for a in (-1, 0, 1) for b in (-1, 0, 1) if a or b]:
                    nr, nc = r + dr, c + dc
                    if (1 <= nr <= ws.max_row and 1 <= nc <= ws.max_column
                            and (nr, nc) not in visited
                            and not _is_empty_value(ws.cell(row=nr, column=nc).value)):
                        queue.append((nr, nc))
            if max_r - min_r >= 1 or max_c - min_c >= 0:
                blocks.append(CellRange(min_r, min_c, max_r, max_c))

    return blocks


# ── Parser principal ───────────────────────────────────────────────────────

def parse_workbook(file_path: str, mapping: dict | None = None) -> WorkbookModel:
    """
    Lit le classeur Excel et construit le WorkbookModel.
    mapping : dict optionnel {sheet_name: {zones: [{name, range}]}}
    """
    wb    = openpyxl.load_workbook(file_path, data_only=True)
    model = WorkbookModel(file_path=str(file_path))

    global_named: dict[str, str] = {}
    for name, defn in wb.defined_names.items():
        dests = list(defn.destinations)
        if dests:
            global_named[name] = dests[0][1]

    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        sheet_model = SheetModel(name=sheet_name, index=idx, page_order=idx)
        sheet_model.dimensions = (ws.max_row or 0, ws.max_column or 0)

        non_empty_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if not _is_empty_value(cell.value):
                    sheet_model.raw_cells[cell.coordinate] = cell.value
                    non_empty_count += 1
        if non_empty_count < 3:
            continue

        # Plages nommées
        for name, ref in global_named.items():
            if sheet_name in ref or "!" not in ref:
                cr = _parse_ref(ref)
                if cr and not any(t.range == cr for t in sheet_model.tables):
                    sheet_model.named_ranges[name] = ref
                    h, d = _extract_table_data(ws, cr)
                    sheet_model.tables.append(
                        TableZone(name=name, range=cr, source="named_range", headers=h, data=d)
                    )

        if mapping and sheet_name in mapping:
            for z in mapping[sheet_name].get("zones", []):
                cr = _parse_ref(z["range"])
                if cr:
                    h, d = _extract_table_data(ws, cr, is_mapping=True)
                    existing = next((t for t in sheet_model.tables if t.name == z["name"]), None)
                    if existing:
                        sheet_model.tables.remove(existing)
                    sheet_model.tables.append(
                        TableZone(name=z["name"], range=cr, source="manual", headers=h, data=d)
                    )
        elif not sheet_model.tables:
            for i, cr in enumerate(_detect_heuristic_tables(ws)):
                h, d = _extract_table_data(ws, cr)
                sheet_model.tables.append(
                    TableZone(name=f"Zone_{i+1}", range=cr, source="heuristic", headers=h, data=d)
                )

        model.sheets.append(sheet_model)

    wb.close()
    return model


def load_mapping(mapping_path: str) -> dict:
    with open(mapping_path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_mapping(mapping: dict, mapping_path: str) -> None:
    with open(mapping_path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)